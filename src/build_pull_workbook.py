"""Generate ciq_pull.xlsx — the Capital IQ (SPG add-in) bulk-pull workbook.

Open the generated file in Excel with the S&P Capital IQ Pro Office add-in loaded, let it refresh
(twice — EPS_PrePrint depends on Earnings), save, then run `python -m src.ingest_ciq_xlsx`.

Formula signatures are copied from the SPG templates in VIC/CapIQ:
  * SPGRANGEV(ticker,"SP_PRICE_CLOSE",start,end,"Options:...")  — one cell spills a (date, close)
    two-column array (SPG_HistoricalMultiple_vs_PriceChart_v1.xlsm!Intermediate!D7).
  * SPGTable(tickerCell, mnemonicCol, periodCol, asOfCol, "Options:...") — one call fills the column
    UNDER the ticker cell, one value per (mnemonic, period, asOf) row
    (SPG_Energy_Estimates_OilandGas_v1.xlsm!Trends!C9: formula in C9, ticker cell AV9, output AV10:AV455).

Block layout (per ticker, 5 data columns + 1 spacer):
  row 5  ticker name            row 7 headers: key | mnemonic | period | as_of | value
  row 8  SPGTable formula (key column)          row 9  value column: CIQ ticker cell (SPGTable arg 1)
  rows 10.. data rows; SPGTable writes results into the value column at those rows.

Periods: ABSOLUTE strings (FQ32021 ...) built from an anchor cell =SPG(tkr,"336831","FQ0",Sdate)
-> "2026FQ2". Templates only ever call SP_EARNINGS_ANNOUNCE_DATE with absolute periods; relative
FQ-n returned #INVALID FUNCTION PARAMETER in the first refresh. --period-mode relative is kept as an option.

Layout is written to data/ciq_pull_layout.json so the ingest script never guesses.
"""
from __future__ import annotations

import argparse
import json

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .paths import BENCHMARKS_CSV, PULL_LAYOUT, PULL_XLSX, UNIVERSE_CSV

FN = "_xll.SNL.Clients.Office.Excel.Functions."
START_FQ = "1Q16"       # first fiscal quarter of history (inclusive); N_QUARTERS derived at build time
PRICE_START = "2015-11-02"  # ~2 months before the first quarter's print so t-1 exists
FIRST_ROW = 10          # first data row in Earnings / EPS_PrePrint / Fallback_Scalar
STRIDE = 6              # columns per ticker block (5 data + 1 spacer)

# mnemonics pulled per historical quarter (as-of = Sdate, i.e. "final" values)
Q_MNEMONICS = [
    ("announce_date", "SP_EARNINGS_ANNOUNCE_DATE"),
    ("period_end", "SP_PERIOD_END"),
    ("fq_label", "336831"),  # numeric item id used in EarningsWatchDashboard -> "2021FQ3"
    ("eps_est", "SP_EPS_EST"),
    ("eps_actual", "SP_EST_ACT_EPS"),
    ("eps_surprise_pct", "SP_EST_EPS_SURPRISE_PERCENT"),
    ("eps_num_est", "SP_EPS_NUM_EST"),
]
# forward quarter rows (FQ+1): (key, mnemonic, asof_kind)
FWD_ROWS = [
    ("next_announce_date", "SP_EARNINGS_ANNOUNCE_DATE", "sdate"),
    ("next_period_end", "SP_PERIOD_END", "sdate"),
    ("next_fq_label", "336831", "sdate"),
    ("next_eps_est_now", "SP_EPS_EST", "sdate"),
    ("next_eps_est_m28d", "SP_EPS_EST", "sdate_m28"),
    ("next_eps_num_est", "SP_EPS_NUM_EST", "sdate"),
]

HDR_FILL = PatternFill("solid", fgColor="DDEBF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
OUT_FILL = PatternFill("solid", fgColor="E2EFDA")
BOLD = Font(bold=True)

SDATE, SDATE_M28, PSTART = "Universe!$B$1", "Universe!$B$2", "Universe!$B$3"


def quarters_since(start_fq: str, today=None) -> int:
    """Number of fiscal quarters from start_fq through the last *completed* calendar quarter (FQ0 proxy)."""
    import datetime as _dt
    q, y = start_fq.upper().split("Q")
    y = int(y); y = y if y > 100 else 2000 + y
    t = today or _dt.date.today()
    last_q_idx = t.year * 4 + (t.month - 1) // 3 - 1          # last completed calendar quarter
    return last_q_idx - (y * 4 + int(q) - 1) + 1


def rel_period(offset: int) -> str:
    return "FQ0" if offset == 0 else f"FQ{offset:+d}"


def abs_period_formula(anchor_cell: str, offset: int) -> str:
    """Anchor label '2026FQ2' (item 336831 at FQ0) -> absolute period 'FQ{q}{yyyy}' shifted by offset quarters."""
    idx = f"(INT(LEFT({anchor_cell},4))*4+RIGHT({anchor_cell},1)-1{offset:+d})"
    return f'=IF(ISERROR(LEFT({anchor_cell},4)*1),"","FQ"&(MOD({idx},4)+1)&INT({idx}/4))'


class Block:
    """Column bookkeeping for one ticker block on a long-form sheet."""

    HEADERS = ("key", "mnemonic", "period", "as_of", "value")
    WIDTHS = (24, 28, 10, 11, 13)

    def __init__(self, ws, col: int, ticker: str, ciq_ticker: str):
        self.ws, self.col, self.ticker, self.ciq = ws, col, ticker, ciq_ticker
        self.kcol, self.mcol, self.pcol, self.acol, self.vcol = (get_column_letter(col + i) for i in range(5))
        ws.cell(5, col, ticker).font = BOLD
        for i, h in enumerate(self.HEADERS):
            c = ws.cell(7, col + i, h)
            c.font, c.fill = BOLD, HDR_FILL
        ws.cell(9, col, "ticker cell (SPGTable arg 1) ->")
        ws.cell(9, col + 4, ciq_ticker).fill = INPUT_FILL
        for i, w in enumerate(self.WIDTHS):
            ws.column_dimensions[get_column_letter(col + i)].width = w
        self.rows = []

    def add_row(self, r: int, key: str, mnemonic: str, period, asof_formula: str, offset: int):
        ws = self.ws
        ws.cell(r, self.col, f"{key}|{offset:+d}")
        ws.cell(r, self.col + 1, mnemonic)
        ws.cell(r, self.col + 2, period)
        ws.cell(r, self.col + 3, asof_formula).number_format = "yyyy-mm-dd"
        ws.cell(r, self.col + 4).fill = OUT_FILL
        self.rows.append({"row": r, "key": key, "offset": offset, "mnemonic": mnemonic})

    def write_table_formula(self, first_row: int, last_row: int):
        f = (f'={FN}SPGTable(${self.vcol}$9,${self.mcol}${first_row}:${self.mcol}${last_row},'
             f'${self.pcol}${first_row}:${self.pcol}${last_row},${self.acol}${first_row}:${self.acol}${last_row},'
             f'"Options:Curr=USD,NA=NA")')
        self.ws.cell(8, self.col, f)
        self.ws.cell(8, self.col + 1, "<- SPGTable call; results land in the value column below the ticker cell")

    def layout(self, first_row: int, last_row: int) -> dict:
        return {"key_col": self.kcol, "mnemonic_col": self.mcol, "period_col": self.pcol, "asof_col": self.acol,
                "value_col": self.vcol, "ticker_cell": f"{self.vcol}9", "first_row": first_row, "last_row": last_row,
                "formula_cell": f"{self.kcol}8", "rows": self.rows}


def build(period_mode: str = "absolute", out=PULL_XLSX, layout_out=PULL_LAYOUT, start_fq: str = START_FQ) -> dict:
    N_QUARTERS = quarters_since(start_fq)
    uni = pd.read_csv(UNIVERSE_CSV)
    bench = pd.read_csv(BENCHMARKS_CSV)
    tickers = list(uni.ticker)
    ciq = dict(zip(uni.ticker, uni.ciq_ticker))
    ciq.update(dict(zip(bench.ticker, bench.ciq_ticker)))
    instruments = tickers + [b for b in bench.ticker if b in set(uni.benchmark)]

    wb = Workbook()
    layout: dict = {"period_mode": period_mode, "n_quarters": N_QUARTERS, "start_fq": start_fq, "sheets": {}}

    # ---------------- Universe / settings ----------------
    ws = wb.active
    ws.title = "Universe"
    ws["A1"], ws["B1"] = "Sdate (as-of, today)", "=TODAY()"
    ws["A2"], ws["B2"] = "Sdate - 28d", "=B1-28"
    y, m, d = (int(x) for x in PRICE_START.split("-"))
    ws["A3"], ws["B3"] = "Price start", f"=DATE({y},{m},{d})"
    ws["A4"], ws["B4"] = "Period mode", period_mode
    for c in ("B1", "B2", "B3", "B4"):
        ws[c].fill = INPUT_FILL
    for c in ("B1", "B2", "B3"):
        ws[c].number_format = "yyyy-mm-dd"
    ws["A6"], ws["B6"], ws["C6"] = "ticker", "ciq_ticker", "role"
    for c in ("A6", "B6", "C6"):
        ws[c].font, ws[c].fill = BOLD, HDR_FILL
    r = 7
    for t in tickers:
        ws.cell(r, 1, t); ws.cell(r, 2, ciq[t]); ws.cell(r, 3, "stock"); r += 1
    for b in bench.ticker:
        ws.cell(r, 1, b); ws.cell(r, 2, ciq[b]); ws.cell(r, 3, "benchmark"); r += 1
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18

    # ---------------- Prices (SPGRANGEV) ----------------
    wp = wb.create_sheet("Prices")
    wp["A1"] = ("One SPGRANGEV per instrument (row 7). The add-in spills dates into the left column and "
                "closes into the formula column from row 8 down.")
    wp["A2"] = f"Verify: ~{int((pd.Timestamp.today()-pd.Timestamp(PRICE_START)).days*252/365)} rows per instrument; closes are split-adjusted (check a known split)."
    price_layout = {}
    col = 2
    for inst in instruments:
        dcol, pcol = get_column_letter(col), get_column_letter(col + 1)
        wp[f"{dcol}5"] = inst
        wp[f"{dcol}5"].font = BOLD
        wp[f"{pcol}5"] = ciq[inst]
        wp[f"{dcol}6"], wp[f"{pcol}6"] = "date", "close"
        wp[f"{pcol}7"] = (f'={FN}SPGRANGEV("{ciq[inst]}","SP_PRICE_CLOSE",{PSTART},{SDATE},'
                          f'"Options: NA=NA(), Sort=Asc,Dates=Before,currency=USD,Caption=Close")')
        wp.column_dimensions[dcol].width = 12
        price_layout[inst] = {"date_col": dcol, "close_col": pcol, "first_row": 8, "formula_cell": f"{pcol}7"}
        col += 3
    layout["sheets"]["Prices"] = price_layout

    # ---------------- Earnings (SPGTable, long form) ----------------
    we = wb.create_sheet("Earnings")
    we["A1"] = ("Per ticker: one SPGTable call (row 8, key column) fills the value column below the ticker cell "
                "(row 9) for every (mnemonic, period, as-of) row. as-of = Sdate except the -28d revision row.")
    we["A2"] = ("Periods are absolute (FQ32021...) derived from the anchor in row 4 (item 336831 at FQ0). "
                "If the anchor is blank/error the period cells stay blank -> refresh again.")
    earn_layout = {}
    col = 2
    for t in tickers:
        b = Block(we, col, t, ciq[t])
        anchor = f"${b.vcol}$4"
        we.cell(4, col, "FQ0 anchor label (item 336831) ->")
        if period_mode == "absolute":
            we.cell(4, col + 4, f'={FN}SPG("{ciq[t]}","336831","FQ0",{SDATE})').fill = INPUT_FILL
        r = FIRST_ROW
        for q in range(-(N_QUARTERS - 1), 1):          # FQ-19 .. FQ0
            for key, mn in Q_MNEMONICS:
                per = abs_period_formula(anchor, q) if period_mode == "absolute" else rel_period(q)
                b.add_row(r, key, mn, per, f"={SDATE}", q)
                r += 1
        for key, mn, asof_kind in FWD_ROWS:
            per = abs_period_formula(anchor, 1) if period_mode == "absolute" else "FQ+1"
            b.add_row(r, key, mn, per, f"={SDATE_M28 if asof_kind == 'sdate_m28' else SDATE}", 1)
            r += 1
        last_row = r - 1
        b.write_table_formula(FIRST_ROW, last_row)
        earn_layout[t] = b.layout(FIRST_ROW, last_row)
        col += STRIDE
    layout["sheets"]["Earnings"] = earn_layout

    # ---------------- EPS_PrePrint (consensus as of announce date - 1) ----------------
    wq = wb.create_sheet("EPS_PrePrint")
    wq["A1"] = ("Point-in-time consensus: SP_EPS_EST with as-of = announce date - 1. period and announce date "
                "reference the Earnings sheet -> refresh twice (or after Earnings is populated).")
    pre_layout = {}
    col = 2
    for t in tickers:
        b = Block(wq, col, t, ciq[t])
        e = earn_layout[t]
        r = FIRST_ROW
        for q in range(-(N_QUARTERS - 1), 1):
            ann_row = next(x["row"] for x in e["rows"] if x["key"] == "announce_date" and x["offset"] == q)
            per = f"=Earnings!${e['period_col']}${ann_row}"
            asof = f'=IF(ISNUMBER(Earnings!${e["value_col"]}${ann_row}),Earnings!${e["value_col"]}${ann_row}-1,"")'
            b.add_row(r, "eps_est_preprint", "SP_EPS_EST", per, asof, q)
            r += 1
        last_row = r - 1
        b.write_table_formula(FIRST_ROW, last_row)
        pre_layout[t] = b.layout(FIRST_ROW, last_row)
        col += STRIDE
    layout["sheets"]["EPS_PrePrint"] = pre_layout

    # ---------------- Fallback: scalar SPG loop (EarningsWatch pattern) ----------------
    wf = wb.create_sheet("Fallback_Scalar")
    wf["A1"] = ("Only used if SPGTable fails: one SPG() per (ticker, quarter, mnemonic), as in "
                "SPG_EarningsWatchDashboard!FocusCompany!AZ8:BF15. Same rows as Earnings; periods reference Earnings.")
    fb_layout = {}
    col = 2
    for t in tickers:
        b = Block(wf, col, t, ciq[t])
        e = earn_layout[t]
        r = FIRST_ROW
        for row in e["rows"]:
            per = f"=Earnings!${e['period_col']}${row['row']}"
            asof = SDATE_M28 if row["key"] == "next_eps_est_m28d" else SDATE
            wf.cell(r, col, f"{row['key']}|{row['offset']:+d}")
            wf.cell(r, col + 1, row["mnemonic"])
            wf.cell(r, col + 2, per)
            wf.cell(r, col + 3, f"={asof}").number_format = "yyyy-mm-dd"
            wf.cell(r, col + 4, f'={FN}SPG(${b.vcol}$9,"{row["mnemonic"]}",{b.pcol}{r},{b.acol}{r},"Options:NA=NA")')
            b.rows.append({"row": r, "key": row["key"], "offset": row["offset"], "mnemonic": row["mnemonic"]})
            r += 1
        wf.cell(8, col, "(scalar SPG per row - no SPGTable)")
        fb_layout[t] = b.layout(FIRST_ROW, r - 1)
        col += STRIDE
    layout["sheets"]["Fallback_Scalar"] = fb_layout

    # ---------------- Test bed: one cell per candidate call ----------------
    wt = wb.create_sheet("Test")
    t0 = tickers[0]
    c0 = ciq[t0]
    wt["A1"] = f"Mnemonic / period test bed for {c0} - instant feedback per cell. Rows 4-6 & 11-16 should work per templates."
    tests = [
        ("336831 @ FQ0 (anchor)", f'={FN}SPG("{c0}","336831","FQ0",{SDATE})'),
        ("SP_EARNINGS_ANNOUNCE_DATE @ absolute FQ32021", f'={FN}SPG("{c0}","SP_EARNINGS_ANNOUNCE_DATE","FQ32021",{SDATE})'),
        ("SP_EARNINGS_ANNOUNCE_DATE @ absolute FQ12023", f'={FN}SPG("{c0}","SP_EARNINGS_ANNOUNCE_DATE","FQ12023",{SDATE})'),
        ("SP_EARNINGS_ANNOUNCE_DATE @ FQ32021, no as-of", f'={FN}SPG("{c0}","SP_EARNINGS_ANNOUNCE_DATE","FQ32021")'),
        ("SP_EARNINGS_ANNOUNCE_DATE @ relative FQ-19 (failed in refresh 1)", f'={FN}SPG("{c0}","SP_EARNINGS_ANNOUNCE_DATE","FQ-19",{SDATE})'),
        ("SP_EARNINGS_ANNOUNCE_DATE @ FQ0", f'={FN}SPG("{c0}","SP_EARNINGS_ANNOUNCE_DATE","FQ0",{SDATE})'),
        ("SP_EARNINGS_ANNOUNCE_DATE @ FQ+1 (next call date)", f'={FN}SPG("{c0}","SP_EARNINGS_ANNOUNCE_DATE","FQ+1",{SDATE})'),
        ("SP_PERIOD_END @ FQ32021", f'={FN}SPG("{c0}","SP_PERIOD_END","FQ32021",{SDATE})'),
        ("SP_EPS_EST @ FQ32021 as-of Sdate (final consensus)", f'={FN}SPG("{c0}","SP_EPS_EST","FQ32021",{SDATE},"Options:NA=NA")'),
        ("SP_EPS_EST @ FQ32021 as-of 2021-10-18 (pre-print)", f'={FN}SPG("{c0}","SP_EPS_EST","FQ32021",DATE(2021,10,18),"Options:NA=NA")'),
        ("SP_EST_ACT_EPS @ FQ32021", f'={FN}SPG("{c0}","SP_EST_ACT_EPS","FQ32021",{SDATE},"Options:NA=NA")'),
        ("SP_EST_EPS_SURPRISE_PERCENT @ FQ32021", f'={FN}SPG("{c0}","SP_EST_EPS_SURPRISE_PERCENT","FQ32021",{SDATE},"Options:NA=NA")'),
        ("SP_EPS_NUM_EST @ FQ32021", f'={FN}SPG("{c0}","SP_EPS_NUM_EST","FQ32021",{SDATE},"Options:NA=NA")'),
        ("SP_PRICE_CLOSE @ 2021-10-19", f'={FN}SPG("{c0}","SP_PRICE_CLOSE",DATE(2021,10,19),"Options:Curr=USD")'),
    ]
    wt["A3"], wt["B3"], wt["C3"] = "test", "result", "formula (text)"
    for c in ("A3", "B3", "C3"):
        wt[c].font, wt[c].fill = BOLD, HDR_FILL
    for i, (name, f) in enumerate(tests):
        wt.cell(4 + i, 1, name)
        wt.cell(4 + i, 2, f)
        wt.cell(4 + i, 3, "'" + f)
    wt.column_dimensions["A"].width = 58
    wt.column_dimensions["B"].width = 18
    wt.column_dimensions["C"].width = 100
    layout["sheets"]["Test"] = {"first_row": 4, "n": len(tests)}

    _backup_existing(out)
    wb.save(out)
    layout_out.parent.mkdir(parents=True, exist_ok=True)
    layout_out.write_text(json.dumps(layout, indent=1))
    return layout


def _backup_existing(path):
    """Never clobber a workbook that may hold refreshed (cached) values: copy it to data/raw first."""
    import shutil, datetime as _dt
    from pathlib import Path
    from .paths import RAW
    path = Path(path)
    if path.exists():
        dst = RAW / f"{path.stem}_{_dt.datetime.now():%Y%m%d_%H%M%S}{path.suffix}"
        shutil.copy2(path, dst)
        print(f"  (backed up existing {path.name} -> {dst})")


def build_etf_only(out=None, layout_out=None) -> dict:
    """Prices-only workbook for the benchmark ETFs (one SPGRANGEV each) -> ciq_pull_etf.xlsx."""
    from .paths import ROOT, DATA
    out = out or ROOT / "ciq_pull_etf.xlsx"
    layout_out = layout_out or DATA / "ciq_pull_etf_layout.json"
    bench = pd.read_csv(BENCHMARKS_CSV)
    wb = Workbook()
    ws = wb.active
    ws.title = "Universe"
    ws["A1"], ws["B1"] = "Sdate (as-of, today)", "=TODAY()"
    y, m, d = (int(x) for x in PRICE_START.split("-"))
    ws["A3"], ws["B3"] = "Price start", f"=DATE({y},{m},{d})"
    ws["B1"].number_format = ws["B3"].number_format = "yyyy-mm-dd"
    wp = wb.create_sheet("Prices")
    wp["A1"] = "Benchmark ETF closes: one SPGRANGEV per ETF (row 7); dates spill left, closes under the formula."
    price_layout, col = {}, 2
    for inst, ciq_t in zip(bench.ticker, bench.ciq_ticker):
        dcol, pcol = get_column_letter(col), get_column_letter(col + 1)
        wp[f"{dcol}5"], wp[f"{pcol}5"] = inst, ciq_t
        wp[f"{dcol}5"].font = BOLD
        wp[f"{dcol}6"], wp[f"{pcol}6"] = "date", "close"
        wp[f"{pcol}7"] = (f'={FN}SPGRANGEV("{ciq_t}","SP_PRICE_CLOSE",{PSTART},{SDATE},'
                          f'"Options: NA=NA(), Sort=Asc,Dates=Before,currency=USD,Caption=Close")')
        wp.column_dimensions[dcol].width = 12
        price_layout[inst] = {"date_col": dcol, "close_col": pcol, "first_row": 8, "formula_cell": f"{pcol}7"}
        col += 3
    layout = {"etf_only": True, "sheets": {"Prices": price_layout}}
    _backup_existing(out)
    wb.save(out)
    layout_out.write_text(json.dumps(layout, indent=1))
    print(f"Wrote {out} ({len(price_layout)} SPGRANGEV calls); layout -> {layout_out}")
    return layout


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--etf-only", action="store_true", help="write ciq_pull_etf.xlsx (benchmark ETF prices only)")
    ap.add_argument("--period-mode", choices=["absolute", "relative"], default="absolute")
    ap.add_argument("--start-fq", default=START_FQ, help="first fiscal quarter, e.g. 1Q16")
    args = ap.parse_args()
    if args.etf_only:
        build_etf_only()
        return
    layout = build(args.period_mode, start_fq=args.start_fq)
    n_t = len(layout["sheets"]["Earnings"])
    n_p = len(layout["sheets"]["Prices"])
    print(f"Wrote {PULL_XLSX}  (period mode: {args.period_mode}, {layout['n_quarters']} quarters from {args.start_fq})\n  Prices: {n_p} SPGRANGEV calls"
          f"\n  Earnings: {n_t} SPGTable calls\n  EPS_PrePrint: {n_t} SPGTable calls\n  layout -> {PULL_LAYOUT}")


if __name__ == "__main__":
    main()
