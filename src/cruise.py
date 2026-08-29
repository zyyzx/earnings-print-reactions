"""Earnings pack for a focus name + peer group (generalized from the OSW/cruise build).

  python -m src.cruise workbook --group <g>            # generate ciq_pull_<g>.xlsx (user refreshes in Excel)
  python -m src.cruise yahoo    --group <g>            # bootstrap prices + EPS + dates/timing from Yahoo
  python -m src.cruise ingest   --group <g>            # read the refreshed CapIQ workbook (adds Rev/EBITDA consensus)
  python -m src.cruise pack     --group <g>            # output/<g>/<g>_earnings_pack.xlsx + one PNG per ticker
  python -m src.cruise charts   --group <g> --focus T  # combined indexed chart, beats panels, vs-peers panel

Config: config/<g>_universe.csv with columns ticker,ciq_ticker,timing_default,fye_month,notes[,role]
(role: focus|peer; default = first row is the focus name, the rest are peers). Start quarter via --start (default 1Q23).
Peer basket for relative math = rows with role=peer, ex-self; the focus name is never in the basket.
"""
from __future__ import annotations

import argparse
import json
from datetime import date

import numpy as np
import pandas as pd

from . import db, labels
from .ingest_ciq_xlsx import as_date, as_num
from .paths import CONFIG, DATA, OUTPUT, ROOT

GROUP = "cruise"
UNIVERSE = CONFIG / "cruise_universe.csv"
PULL = ROOT / "ciq_pull_cruise.xlsx"
LAYOUT = DATA / "ciq_pull_cruise_layout.json"
OUT = OUTPUT / "cruise"
START_FQ = "1Q23"


def set_group(group: str, start_fq: str | None = None):
    """Point module paths at a named group (config/<group>_universe.csv etc.)."""
    global GROUP, UNIVERSE, PULL, LAYOUT, OUT, START_FQ
    GROUP = group
    UNIVERSE = CONFIG / f"{group}_universe.csv"
    PULL = ROOT / f"ciq_pull_{group}.xlsx"
    LAYOUT = DATA / f"ciq_pull_{group}_layout.json"
    OUT = OUTPUT / group
    if start_fq:
        START_FQ = start_fq


def load_universe() -> pd.DataFrame:
    u = pd.read_csv(UNIVERSE)
    if "role" not in u.columns:
        u["role"] = ["focus"] + ["peer"] * (len(u) - 1)
    return u
PRICE_START = "2022-12-15"
FN = "_xll.SNL.Clients.Office.Excel.Functions."
FIRST_ROW = 10

Q_MNEMONICS = [
    ("announce_date", "SP_EARNINGS_ANNOUNCE_DATE"),
    ("period_end", "SP_PERIOD_END"),
    ("fq_label", "336831"),
    ("rev_est", "SP_REV_EST"),
    ("rev_actual", "SP_EST_ACT_REV"),
    ("rev_surprise_pct", "SP_EST_REV_SURPRISE_PERCENT"),
    ("ebitda_est", "SP_EBITDA_EST"),
    ("ebitda_actual", "SP_EST_ACT_EBITDA"),
    ("eps_est", "SP_EPS_EST"),
    ("eps_actual", "SP_EST_ACT_EPS"),
    ("eps_surprise_pct", "SP_EST_EPS_SURPRISE_PERCENT"),
]

SCHEMA = """
CREATE TABLE IF NOT EXISTS cruise_earnings (
    ticker TEXT NOT NULL, fq_label TEXT NOT NULL, announce_date TEXT, period_end TEXT, timing TEXT,
    rev_est REAL, rev_actual REAL, rev_surprise_pct REAL,
    ebitda_est REAL, ebitda_actual REAL,
    eps_est REAL, eps_actual REAL, eps_surprise_pct REAL,
    source TEXT, PRIMARY KEY (ticker, fq_label));
"""


def n_quarters() -> int:
    q, y = START_FQ.upper().split("Q")[0], START_FQ.split("Q")[1]
    q, y = int(q), 2000 + int(y) if int(y) < 100 else int(y)
    t = date.today()
    return (t.year * 4 + (t.month - 1) // 3 - 1) - (y * 4 + q - 1) + 1


def make_workbook():
    """One SPGTable per ticker (absolute periods off a 336831 anchor) + one SPGRANGEV per ticker for prices."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    uni = load_universe()
    NQ = n_quarters()
    wb = Workbook()
    ws = wb.active
    ws.title = "Universe"
    ws["A1"], ws["B1"] = "Sdate", "=TODAY()"
    y, m, d = (int(x) for x in PRICE_START.split("-"))
    ws["A2"], ws["B2"] = "Price start", f"=DATE({y},{m},{d})"
    ws["B1"].number_format = ws["B2"].number_format = "yyyy-mm-dd"
    SDATE = "Universe!$B$1"

    wp = wb.create_sheet("Prices")
    wp["A1"] = "One SPGRANGEV per ticker (row 7): dates spill left of the formula column, closes below it."
    price_layout, col = {}, 2
    for r in uni.itertuples():
        dcol, pcol = get_column_letter(col), get_column_letter(col + 1)
        wp[f"{dcol}5"] = r.ticker
        wp[f"{dcol}5"].font = Font(bold=True)
        wp[f"{pcol}5"] = r.ciq_ticker
        wp[f"{dcol}6"], wp[f"{pcol}6"] = "date", "close"
        wp[f"{pcol}7"] = (f'={FN}SPGRANGEV("{r.ciq_ticker}","SP_PRICE_CLOSE",Universe!$B$2,{SDATE},'
                          f'"Options: NA=NA(), Sort=Asc,Dates=Before,currency=USD,Caption=Close")')
        price_layout[r.ticker] = {"date_col": dcol, "close_col": pcol, "first_row": 8}
        col += 3

    we = wb.create_sheet("Earnings")
    we["A1"] = ("Per ticker: one SPGTable call (row 8, key column) fills the value column under the ticker cell (row 9). "
                "Revenue/EBITDA in USD millions. Periods are absolute, derived from the 336831 anchor in row 4.")
    earn_layout, col = {}, 2
    for r in uni.itertuples():
        kcol, mcol, pcol, acol, vcol = (get_column_letter(col + i) for i in range(5))
        we.cell(4, col, "FQ0 anchor ->")
        we.cell(4, col + 4, f'={FN}SPG("{r.ciq_ticker}","336831","FQ0",{SDATE})').fill = PatternFill("solid", fgColor="FFF2CC")
        we.cell(5, col, r.ticker).font = Font(bold=True)
        for i, h in enumerate(("key", "mnemonic", "period", "as_of", "value")):
            we.cell(7, col + i, h).font = Font(bold=True)
        we.cell(9, col, "ticker cell ->")
        we.cell(9, col + 4, r.ciq_ticker).fill = PatternFill("solid", fgColor="FFF2CC")
        anchor = f"${vcol}$4"
        rows, rr = [], FIRST_ROW
        for q in range(-(NQ - 1), 1):
            idx = f"(INT(LEFT({anchor},4))*4+RIGHT({anchor},1)-1{q:+d})"
            per = f'=IF(ISERROR(LEFT({anchor},4)*1),"","FQ"&(MOD({idx},4)+1)&INT({idx}/4))'
            for key, mn in Q_MNEMONICS:
                we.cell(rr, col, f"{key}|{q:+d}")
                we.cell(rr, col + 1, mn)
                we.cell(rr, col + 2, per)
                we.cell(rr, col + 3, f"={SDATE}").number_format = "yyyy-mm-dd"
                rows.append({"row": rr, "key": key, "offset": q})
                rr += 1
        last = rr - 1
        we.cell(8, col, (f'={FN}SPGTable(${vcol}$9,${mcol}${FIRST_ROW}:${mcol}${last},'
                         f'${pcol}${FIRST_ROW}:${pcol}${last},${acol}${FIRST_ROW}:${acol}${last},'
                         f'"Options:Curr=USD,Mag=Millions,NA=NA")'))
        earn_layout[r.ticker] = {"value_col": vcol, "rows": rows}
        col += 6
    wb.save(PULL)
    LAYOUT.write_text(json.dumps({"n_quarters": NQ, "sheets": {"Prices": price_layout, "Earnings": earn_layout}}, indent=1))
    print(f"Wrote {PULL} ({len(uni)} tickers x {NQ} quarters; refresh in Excel, then: python -m src.cruise ingest)")


def yahoo():
    import yfinance as yf
    uni = load_universe()
    con = db.connect()
    con.executescript(SCHEMA)
    for r in uni.itertuples():
        h = yf.Ticker(r.ticker).history(start="2022-12-01", auto_adjust=False)
        idx = pd.to_datetime(h.index).tz_localize(None).normalize()
        px = pd.DataFrame({"ticker": r.ticker, "date": idx, "close": h["Adj Close"].values, "source": "yahoo"})
        con.execute("DELETE FROM prices WHERE ticker=? AND source='yahoo'", (r.ticker,))
        db.upsert(con, "prices", px, ["ticker", "date"])
        ed = yf.Ticker(r.ticker).get_earnings_dates(limit=30)
        recs = []
        if ed is not None and not ed.empty:
            ed = ed[~ed.index.duplicated(keep="first")].sort_index()
            ts = ed.index.tz_convert("US/Eastern") if ed.index.tz is not None else ed.index
            for t_, (_, row) in zip(ts, ed.iterrows()):
                ann = pd.Timestamp(t_.date())
                if ann >= pd.Timestamp.today().normalize() or pd.isna(row.get("Reported EPS")):
                    continue
                lab = labels.label_from_date_guess(ann, int(r.fye_month))
                if lab is None or labels.sort_key(lab) < labels.sort_key(START_FQ):
                    continue
                recs.append(dict(ticker=r.ticker, fq_label=lab, announce_date=ann.strftime("%Y-%m-%d"), period_end=None,
                                 timing="BMO" if t_.hour < 12 else "AMC",
                                 rev_est=None, rev_actual=None, rev_surprise_pct=None, ebitda_est=None, ebitda_actual=None,
                                 eps_est=as_num(row.get("EPS Estimate")), eps_actual=as_num(row.get("Reported EPS")),
                                 eps_surprise_pct=as_num(row.get("Surprise(%)")), source="yahoo"))
        if recs:
            db.upsert(con, "cruise_earnings", pd.DataFrame(recs), ["ticker", "fq_label"])
        print(f"  {r.ticker}: {len(px)} closes, {len(recs)} prints")
    print("Done ->", db.DB_PATH)


def ingest():
    from openpyxl import load_workbook
    L = json.loads(LAYOUT.read_text())
    wb = load_workbook(PULL, data_only=True, read_only=False)
    con = db.connect()
    con.executescript(SCHEMA)
    uni = load_universe().set_index("ticker")
    ws = wb["Prices"]
    for t, pl in L["sheets"]["Prices"].items():
        recs, rr = [], pl["first_row"]
        while True:
            d_, p_ = ws[f"{pl['date_col']}{rr}"].value, ws[f"{pl['close_col']}{rr}"].value
            if d_ is None and p_ is None:
                break
            dd, pp = as_date(d_), as_num(p_)
            if dd is not None and pp is not None:
                recs.append((t, dd.strftime("%Y-%m-%d"), pp, "ciq"))
            rr += 1
        if len(recs) > 100:
            con.execute("DELETE FROM prices WHERE ticker=? AND source<>'ciq'", (t,))
            db.upsert(con, "prices", pd.DataFrame(recs, columns=["ticker", "date", "close", "source"]), ["ticker", "date"])
        print(f"  prices {t}: {len(recs)} rows {'OK' if len(recs) > 100 else 'CHECK (not refreshed?)'}")
    ws = wb["Earnings"]
    for t, el in L["sheets"]["Earnings"].items():
        vcol = el["value_col"]
        raw = {}
        for row in el["rows"]:
            raw.setdefault(row["offset"], {})[row["key"]] = ws[f"{vcol}{row['row']}"].value
        fye = int(uni.loc[t, "fye_month"])
        timing = {r[0]: r[1] for r in con.execute("SELECT fq_label, timing FROM cruise_earnings WHERE ticker=? AND timing IS NOT NULL", (t,))}
        recs = []
        for q, vals in raw.items():
            ann, pe = as_date(vals.get("announce_date")), as_date(vals.get("period_end"))
            lab = labels.normalise(vals.get("fq_label")) or labels.from_period_end(pe, fye)
            if ann is None or lab is None or labels.sort_key(lab) < labels.sort_key(START_FQ):
                continue
            recs.append(dict(ticker=t, fq_label=lab, announce_date=ann.strftime("%Y-%m-%d"),
                             period_end=pe.strftime("%Y-%m-%d") if pe is not None else None,
                             timing=timing.get(lab, uni.loc[t, "timing_default"]),
                             rev_est=as_num(vals.get("rev_est")), rev_actual=as_num(vals.get("rev_actual")),
                             rev_surprise_pct=as_num(vals.get("rev_surprise_pct")),
                             ebitda_est=as_num(vals.get("ebitda_est")), ebitda_actual=as_num(vals.get("ebitda_actual")),
                             eps_est=as_num(vals.get("eps_est")), eps_actual=as_num(vals.get("eps_actual")),
                             eps_surprise_pct=as_num(vals.get("eps_surprise_pct")), source="ciq"))
        if recs:
            con.execute("DELETE FROM cruise_earnings WHERE ticker=? AND source<>'ciq'", (t,))
            db.upsert(con, "cruise_earnings", pd.DataFrame(recs), ["ticker", "fq_label"])
        print(f"  earnings {t}: {len(recs)} quarters {'OK' if len(recs) >= 8 else 'CHECK'}")
    print("Done ->", db.DB_PATH)


def peer_list() -> list[str]:
    u = load_universe()
    return list(u.loc[u.role == "peer", "ticker"])


def _events(con, t: str) -> pd.DataFrame:
    e = pd.read_sql("SELECT * FROM cruise_earnings WHERE ticker=? ORDER BY announce_date", con, params=[t])
    if e.empty:
        return e
    px = db.read_prices(con, [t]).set_index("date")["close"]
    ret = px.pct_change()
    peers = [o for o in peer_list() if o != t]
    pw = db.read_prices(con, peers).pivot(index="date", columns="ticker", values="close")
    peer_ret = pw.pct_change().mean(axis=1, skipna=True)
    cal = px.index
    rows = []
    for r in e.itertuples():
        ann = pd.Timestamp(r.announce_date)
        pos = cal.searchsorted(ann, side="right" if (r.timing or "BMO").upper() == "AMC" else "left")
        t0 = cal[pos] if pos < len(cal) else pd.NaT
        d = dict(r._asdict()); d.pop("Index", None)
        d["t0"] = t0
        d["t0_close"] = px.get(t0, np.nan)
        d["t0_pct"] = ret.get(t0, np.nan) * 100 if not pd.isna(t0) else np.nan
        d["t0_peers_pct"] = peer_ret.get(t0, np.nan) * 100 if not pd.isna(t0) else np.nan
        d["t0_rel_pp"] = d["t0_pct"] - d["t0_peers_pct"]
        nxt = cal[pos + 1] if pos + 1 < len(cal) else pd.NaT
        d["tp1_pct"] = ret.get(nxt, np.nan) * 100 if not pd.isna(nxt) else np.nan
        for base in ("rev", "ebitda", "eps"):
            est, act = d.get(f"{base}_est"), d.get(f"{base}_actual")
            if d.get(f"{base}_surprise_pct") is None and est not in (None, 0) and act is not None and not pd.isna(est) and not pd.isna(act):
                d[f"{base}_surprise_pct"] = (act - est) / abs(est) * 100
        rows.append(d)
    df = pd.DataFrame(rows)
    df["sort"] = df.fq_label.map(labels.sort_key)
    return df.sort_values("sort").drop(columns="sort").reset_index(drop=True)


def pack():
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, ScatterChart, Reference, Series
    from openpyxl.chart.marker import Marker
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    OUT.mkdir(parents=True, exist_ok=True)
    uni = load_universe()
    con = db.connect()
    con.executescript(SCHEMA)
    wb = Workbook()
    summary_rows = []
    HDR = PatternFill("solid", fgColor="DDEBF7")
    first = True
    for u in uni.itertuples():
        t = u.ticker
        ev = _events(con, t)
        px = db.read_prices(con, [t])
        px = px[px.date >= "2023-01-01"]
        if ev.empty or px.empty:
            print(f"  {t}: no data, skipped")
            continue
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = t
        ws["A1"] = f"{t} ({u.notes}) - consensus vs results and the stock reaction, 1Q23 onward"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = "Revenue and EBITDA in $M. vs peers = print-day move minus equal-weight cruise operators (ex-self) that day, in points. Surprise = (actual - consensus) / |consensus|. Print-day = first session reflecting the release. Source: S&P Capital IQ (Yahoo where consensus not yet refreshed)."
        ws["A2"].font = Font(size=9, italic=True)
        heads = ["Quarter", "Announced", "Timing", "Rev cons", "Rev actual", "Rev spr %", "EBITDA cons", "EBITDA actual", "EBITDA spr %",
                 "EPS cons", "EPS actual", "EPS spr %", "Print-day %", "Peers that day %", "vs peers (pp)", "Day-after %"]
        for j, h in enumerate(heads, start=1):
            c = ws.cell(4, j, h)
            c.font, c.fill = Font(bold=True), HDR
        for i, r in enumerate(ev.itertuples(), start=5):
            eb_spr = r.ebitda_surprise_pct if hasattr(r, "ebitda_surprise_pct") else None
            vals = [r.fq_label, r.announce_date, r.timing, r.rev_est, r.rev_actual, r.rev_surprise_pct,
                    r.ebitda_est, r.ebitda_actual, eb_spr, r.eps_est, r.eps_actual, r.eps_surprise_pct,
                    r.t0_pct, r.t0_peers_pct, r.t0_rel_pp, r.tp1_pct]
            for j, v in enumerate(vals, start=1):
                if isinstance(v, float) and pd.isna(v):
                    v = None
                c = ws.cell(4 + i - 4, j, v)
                if j in (4, 5, 7, 8):
                    c.number_format = "#,##0"
                elif j in (10, 11):
                    c.number_format = "0.00"
                elif j in (6, 9, 12, 13, 14, 15, 16):
                    c.number_format = "+0.0;-0.0"
                    if isinstance(v, (int, float)):
                        c.font = Font(color="2A78D6" if v > 0 else "E34948")
        for j, w in enumerate([9, 11, 7, 10, 10, 9, 11, 12, 10, 9, 9, 9, 11, 13, 12, 11], start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
        # price data block (right of the table) + chart
        pc0 = len(heads) + 2                              # first price col
        dcol, ccol, mcol = (get_column_letter(pc0 + i) for i in range(3))
        ws.cell(4, pc0, "date").fill = HDR
        ws.cell(4, pc0 + 1, "close").fill = HDR
        ws.cell(4, pc0 + 2, "print").fill = HDR
        t0set = {pd.Timestamp(x).strftime("%Y-%m-%d") for x in ev.t0.dropna()}
        for i, r in enumerate(px.itertuples(), start=5):
            ds = pd.Timestamp(r.date).strftime("%Y-%m-%d")
            ws.cell(i, pc0, r.date).number_format = "yyyy-mm-dd"
            ws.cell(i, pc0 + 1, round(r.close, 2))
            if ds in t0set:
                ws.cell(i, pc0 + 2, round(r.close, 2))
        nrows = len(px)
        ch = ScatterChart()
        ch.title = f"{t} close since 2023; markers = earnings prints"
        ch.style = 2
        ch.height, ch.width = 9.5, 21
        ch.x_axis.number_format = "mmm-yy"
        ch.x_axis.majorTimeUnit = "months"
        ch.legend.position = "b"
        xref = Reference(ws, min_col=pc0, min_row=5, max_row=4 + nrows)
        s1 = Series(Reference(ws, min_col=pc0 + 1, min_row=4, max_row=4 + nrows), xref, title_from_data=True)
        s1.marker = Marker(symbol="none")
        s1.graphicalProperties.line.solidFill = "2A78D6"
        s1.graphicalProperties.line.width = 14000
        ch.series.append(s1)
        s2 = Series(Reference(ws, min_col=pc0 + 2, min_row=4, max_row=4 + nrows), xref, title_from_data=True)
        s2.marker = Marker(symbol="circle", size=7)
        s2.marker.graphicalProperties.solidFill = "E34948"
        s2.graphicalProperties.noFill = True
        ch.series.append(s2)
        ws.add_chart(ch, f"A{len(ev) + 8}")
        # PNG
        fig, ax = plt.subplots(figsize=(11, 4.5), dpi=110)
        ax.plot(px.date, px.close, color="#2a78d6", lw=1.3, label="close")
        up = ev[ev.t0_pct > 0]; dn = ev[ev.t0_pct <= 0]
        ax.scatter(up.t0, up.t0_close, color="#1baf7a", zorder=3, s=42, label="print, up on the day")
        ax.scatter(dn.t0, dn.t0_close, color="#e34948", zorder=3, s=42, label="print, down on the day")
        for r in ev.itertuples():
            if not pd.isna(r.t0):
                ax.annotate(r.fq_label, (r.t0, r.t0_close), textcoords="offset points", xytext=(0, 9),
                            fontsize=7, ha="center", color="#52514e")
        ax.set_title(f"{t} ({u.notes}) since 2023; earnings prints marked, colored by print-day move")
        ax.legend(frameon=False, fontsize=8, loc="upper left")
        ax.grid(axis="y", color="#e6e4df", lw=0.6)
        for sp in ("top", "right"):
            ax.spines[sp].set_visible(False)
        fig.tight_layout()
        fig.savefig(OUT / f"{t}.png")
        plt.close(fig)
        beats = int((ev.eps_surprise_pct > 0).sum()); n = int(ev.eps_surprise_pct.notna().sum())
        summary_rows.append([t, str(u.notes), len(ev), f"{beats}/{n}",
                             round(float(ev.t0_pct.abs().mean()), 1), round(float(ev.t0_pct.mean()), 1),
                             ev.fq_label.iloc[0], ev.fq_label.iloc[-1]])
        print(f"  {t}: {len(ev)} prints -> sheet + {t}.png")
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = f"{GROUP} group: earnings prints since {START_FQ}"
    ws["A1"].font = Font(bold=True, size=13)
    for j, h in enumerate(["Ticker", "Name", "Prints", "EPS beats", "Avg |print-day| %", "Avg print-day %", "First", "Last"], start=1):
        c = ws.cell(3, j, h)
        c.font, c.fill = Font(bold=True), HDR
    for i, row in enumerate(summary_rows, start=4):
        for j, v in enumerate(row, start=1):
            ws.cell(i, j, v)
    for j, w in enumerate([8, 22, 8, 10, 16, 15, 8, 8], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    out = OUT / f"{GROUP}_earnings_pack.xlsx"
    wb.save(out)
    print(f"pack -> {out}")


def charts(focus: str | None = None):
    """Combined indexed price chart, per-ticker beats panels, and the focus-vs-peers print-day panel."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates

    OUT.mkdir(parents=True, exist_ok=True)
    uni = load_universe()
    focus = focus or uni.loc[uni.role == "focus", "ticker"].iloc[0]
    tickers = list(uni.ticker)
    peers = [t for t in tickers if t != focus]
    con = db.connect()
    PALETTE = ["#eb6834", "#1baf7a", "#eda100", "#c65fa0", "#4a3aa7", "#008300", "#e34948", "#9a9892"]
    colors = {focus: "#2a78d6", **{t: PALETTE[i % len(PALETTE)] for i, t in enumerate(peers)}}
    names = dict(zip(uni.ticker, uni.notes.fillna(uni.ticker)))
    base = pd.Timestamp(f"20{START_FQ.split('Q')[1]}-01-01")

    # --- combined indexed
    fig, ax = plt.subplots(figsize=(13, 6.2), dpi=130)
    end_pts = []
    for t in peers + [focus]:
        px = db.read_prices(con, [t])
        px = px[px.date >= base].reset_index(drop=True)
        if px.empty:
            continue
        idx = px.close / px.close.iloc[0] * 100
        f = t == focus
        ax.plot(px.date, idx, color=colors[t], lw=2.4 if f else 1.1, alpha=1.0 if f else 0.55, zorder=5 if f else 2)
        ev = _events(con, t).dropna(subset=["t0"])
        ys = [idx.iloc[px.index[px.date == d][0]] if (px.date == d).any() else np.nan for d in ev.t0]
        ax.scatter(ev.t0, ys, color=colors[t], s=26 if f else 12, zorder=6 if f else 3,
                   alpha=1.0 if f else 0.55, edgecolors="white", linewidths=0.6)
        end_pts.append((px.date.iloc[-1], idx.iloc[-1], t))
    for d_, v, t in sorted(end_pts, key=lambda x: -x[1]):
        ax.annotate(f"{t} {v:,.0f}", (d_, v), textcoords="offset points", xytext=(8, 0), fontsize=9,
                    color=colors[t], weight="bold" if t == focus else "normal", va="center")
    ax.set_title(f"{focus} vs peers, indexed to 100 at {base.date()} (later listings at first close); dots = earnings prints",
                 fontsize=12.5, color="#333")
    ax.grid(axis="y", color="#e6e4df", lw=0.6)
    for sp in ("top", "right"):
        ax.spines[sp].set_visible(False)
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b-%y"))
    handles = [plt.Line2D([0], [0], color=colors[t], lw=2.4 if t == focus else 1.4,
                          alpha=1 if t == focus else 0.6, label=f"{t} ({names[t]})") for t in [focus] + peers]
    ax.legend(handles=handles, frameon=False, fontsize=8.5, loc="upper left")
    fig.tight_layout()
    fig.savefig(OUT / "combined_indexed.png")
    plt.close(fig)

    # --- beats panels (every ticker with EBITDA or revenue consensus)
    for t in tickers:
        ev = _events(con, t)
        if ev.empty or ev.ebitda_est.isna().all():
            continue
        ev = ev.reset_index(drop=True)
        ev["ebitda_spr"] = (ev.ebitda_actual - ev.ebitda_est) / ev.ebitda_est.abs() * 100
        x = np.arange(len(ev))
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 7), dpi=130, sharex=True,
                                       gridspec_kw=dict(height_ratios=[1.25, 1.0], hspace=0.10))
        w = 0.38
        ax1.bar(x - w / 2, ev.rev_surprise_pct, w, color="#9dc3e6", label="Revenue beat (%)")
        ax1.bar(x + w / 2, ev.ebitda_spr, w, color="#2a78d6", label="EBITDA beat (%)")
        ax1.axhline(0, color="#c9c7c1", lw=0.8)
        for xi, v in zip(x, ev.ebitda_spr):
            if not np.isnan(v):
                ax1.annotate(f"{v:+.0f}%", (xi + w / 2, v), textcoords="offset points",
                             xytext=(0, 3 if v >= 0 else -11), ha="center", fontsize=7.5, color="#1F4E79")
        ax1.set_ylabel("beat vs consensus (%)")
        ax1.legend(frameon=False, fontsize=9, loc="upper center", ncol=2)
        e1, e2 = ev.ebitda_spr.iloc[:4].mean(), ev.ebitda_spr.iloc[-4:].mean()
        ax1.set_title(f"{t}: EBITDA beats averaged {e1:+.0f}% in the first year of the sample, {e2:+.0f}% over the last four prints",
                      fontsize=12.5, pad=10)
        ax1.grid(axis="y", color="#e6e4df", lw=0.6); ax1.set_axisbelow(True)
        cols = ["#2a78d6" if v > 0 else "#e34948" for v in ev.t0_pct]
        ax2.bar(x, ev.t0_pct, 0.55, color=cols)
        ax2.axhline(0, color="#c9c7c1", lw=0.8)
        for xi, v in zip(x, ev.t0_pct):
            if not np.isnan(v):
                ax2.annotate(f"{v:+.1f}", (xi, v), textcoords="offset points",
                             xytext=(0, 3 if v >= 0 else -11), ha="center", fontsize=7.5, color="#52514e")
        ax2.set_ylabel("print-day stock move (%)")
        ax2.set_xticks(x, ev.fq_label, fontsize=9)
        ax2.grid(axis="y", color="#e6e4df", lw=0.6); ax2.set_axisbelow(True)
        for ax_ in (ax1, ax2):
            for sp in ("top", "right"):
                ax_.spines[sp].set_visible(False)
        fig.subplots_adjust(left=0.065, right=0.985, top=0.93, bottom=0.10)
        fig.text(0.065, 0.015, "Beat = (actual - consensus) / consensus, S&P Capital IQ. Print-day move = close-to-close on the first session reflecting the release.",
                 fontsize=7.5, color="#52514e")
        fig.savefig(OUT / f"{t}_beats.png")
        plt.close(fig)

    # --- focus vs peers print-day panel
    qs = []
    evs = {}
    for t in tickers:
        e = _events(con, t).dropna(subset=["t0_rel_pp"])
        evs[t] = e
        qs.extend(e.fq_label)
    quarters = sorted(set(qs), key=labels.sort_key)
    qi = {q: i for i, q in enumerate(quarters)}
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12.5, 7.6), dpi=130, sharex=True,
                                   gridspec_kw=dict(height_ratios=[1.0, 1.0], hspace=0.14))
    ev = evs[focus]
    x = [qi[q] for q in ev.fq_label]
    cols = ["#2a78d6" if v > 0 else "#e34948" for v in ev.t0_rel_pp]
    ax1.bar(x, ev.t0_rel_pp, 0.55, color=cols)
    for xi, v in zip(x, ev.t0_rel_pp):
        ax1.annotate(f"{v:+.1f}", (xi, v), textcoords="offset points", xytext=(0, 3 if v >= 0 else -11),
                     ha="center", fontsize=7.6, color="#52514e")
    d = ev.t0_rel_pp
    ax1.set_title(f"{focus} print-day move vs its peer basket: avg {d.mean():+.1f}pp, up {int((d > 0).sum())}/{len(d)} prints",
                  fontsize=11.5, loc="left")
    w = min(0.8 / max(len(peers), 1), 0.22)
    for k, t in enumerate(peers):
        e = evs[t]
        xs = [qi[q] + (k - (len(peers) - 1) / 2) * w for q in e.fq_label]
        ax2.bar(xs, e.t0_rel_pp, w, color=colors[t], label=f"{t} ({names[t]})", alpha=0.92)
    ax2.set_title("The peers, same measure (each vs the other peers that day)", fontsize=11.5, loc="left")
    ax2.legend(frameon=False, fontsize=8.5, ncol=min(len(peers), 4), loc="upper right")
    ax2.set_xticks(range(len(quarters)), quarters, fontsize=9)
    for ax_ in (ax1, ax2):
        ax_.axhline(0, color="#c9c7c1", lw=0.8)
        ax_.grid(axis="y", color="#e6e4df", lw=0.6); ax_.set_axisbelow(True)
        ax_.set_ylabel("vs peers (pp)")
        for sp in ("top", "right"):
            ax_.spines[sp].set_visible(False)
    fig.suptitle("Print-day move minus equal-weight peer basket the same day (percentage points)", fontsize=12.5, y=0.985)
    fig.subplots_adjust(left=0.055, right=0.99, top=0.90, bottom=0.075)
    fig.savefig(OUT / "print_day_vs_peers.png")
    plt.close(fig)
    print(f"charts -> {OUT} (combined_indexed, <T>_beats, print_day_vs_peers)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("cmd", choices=["workbook", "yahoo", "ingest", "pack", "charts"])
    ap.add_argument("--group", default="cruise")
    ap.add_argument("--start", default=None, help="first fiscal quarter, e.g. 1Q23")
    ap.add_argument("--focus", default=None, help="focus ticker for charts (default: role=focus row)")
    a = ap.parse_args()
    set_group(a.group, a.start)
    if a.cmd == "charts":
        charts(a.focus)
    else:
        {"workbook": make_workbook, "yahoo": yahoo, "ingest": ingest, "pack": pack}[a.cmd]()
