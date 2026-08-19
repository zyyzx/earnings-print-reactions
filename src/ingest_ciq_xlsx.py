"""Ingest the refreshed ciq_pull.xlsx (cached values) into data/db.sqlite.

Usage:  python -m src.ingest_ciq_xlsx [--xlsx path] [--use-fallback]
Reports unresolved cells (#PEND / NA / errors) per block so you know what to re-refresh.
"""
from __future__ import annotations

import argparse
import json
from datetime import datetime, date

import pandas as pd
from openpyxl import load_workbook

from . import db, labels
from .paths import PULL_LAYOUT, PULL_XLSX, UNIVERSE_CSV

MISSING_STRINGS = {"", "NA", "N/A", "#PEND", "#N/A", "--", "-", "SPGTABLE", "#NAME?", "#VALUE!"}


def is_missing(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        s = v.strip().upper()
        return s in MISSING_STRINGS or s.startswith("#") or s.startswith("SPG")
    if isinstance(v, float) and pd.isna(v):
        return True
    return False


def as_date(v):
    if is_missing(v):
        return None
    if isinstance(v, (datetime, date)):
        return pd.Timestamp(v).normalize()
    if isinstance(v, (int, float)) and 20000 < v < 80000:      # excel serial
        return pd.Timestamp("1899-12-30") + pd.Timedelta(days=float(v))
    try:
        return pd.Timestamp(str(v)).normalize()
    except Exception:
        return None


def as_num(v):
    if is_missing(v):
        return None
    if isinstance(v, (datetime, date)):
        return None
    try:
        return float(v)
    except Exception:
        return None


def ingest_prices(ws, layout: dict, con) -> None:
    for inst, L in layout.items():
        dcol, pcol, r0 = L["date_col"], L["close_col"], L["first_row"]
        recs, missing = [], 0
        r = r0
        while True:
            d, p = ws[f"{dcol}{r}"].value, ws[f"{pcol}{r}"].value
            if d is None and p is None:
                break
            dd, pp = as_date(d), as_num(p)
            if dd is None or pp is None:
                missing += 1
            else:
                recs.append((inst, dd.strftime("%Y-%m-%d"), pp, "ciq"))
            r += 1
        df = pd.DataFrame(recs, columns=["ticker", "date", "close", "source"])
        if len(df) > 200:   # CapIQ is authoritative: drop other-source rows for this instrument (no mixed adj/unadj series)
            con.execute("DELETE FROM prices WHERE ticker=? AND source<>'ciq'", (inst,))
        n = db.upsert(con, "prices", df, ["ticker", "date"])
        db.log(con, "ciq_pull.xlsx", "prices", n, missing, inst)
        status = "OK" if n > 200 else "CHECK - few/no rows (add-in not refreshed?)"
        print(f"  prices {inst:6s}: {n:5d} rows, {missing} unresolved   {status}")


def read_block(ws, block: dict) -> dict:
    """Return {(key, offset): raw value} for one ticker block."""
    out = {}
    vcol = block["value_col"]
    for row in block["rows"]:
        out[(row["key"], row["offset"])] = ws[f"{vcol}{row['row']}"].value
    return out


def ingest_earnings(wb, layout: dict, con, use_fallback: bool) -> None:
    uni = pd.read_csv(UNIVERSE_CSV).set_index("ticker")
    sheet = "Fallback_Scalar" if use_fallback else "Earnings"
    ws = wb[sheet]
    ws_pre = wb["EPS_PrePrint"] if "EPS_PrePrint" in wb.sheetnames else None
    pre_layout = layout["sheets"].get("EPS_PrePrint", {})
    for t, block in layout["sheets"][sheet].items():
        raw = read_block(ws, block)
        pre = read_block(ws_pre, pre_layout[t]) if (ws_pre is not None and t in pre_layout) else {}
        fye = int(uni.loc[t, "fye_month"]) if t in uni.index else 12
        timing_default = uni.loc[t, "timing_default"] if t in uni.index else None
        # keep per-event timing already learned from a source that has it (Yahoo timestamps)
        known_timing = dict(con.execute("SELECT fq_label, timing FROM earnings WHERE ticker=? AND timing IS NOT NULL", (t,)))
        recs, missing = [], 0
        for q in range(-(layout["n_quarters"] - 1), 1):
            ann = as_date(raw.get(("announce_date", q)))
            pe = as_date(raw.get(("period_end", q)))
            lab = labels.normalise(raw.get(("fq_label", q))) or labels.from_period_end(pe, fye) \
                or (labels.label_from_date_guess(ann, fye) if ann is not None else None)
            if ann is None or lab is None:
                missing += 1
                continue
            recs.append(dict(
                ticker=t, fq_label=lab, offset=q,
                announce_date=ann.strftime("%Y-%m-%d"),
                period_end=pe.strftime("%Y-%m-%d") if pe is not None else None,
                eps_est=as_num(raw.get(("eps_est", q))),
                eps_est_preprint=as_num(pre.get(("eps_est_preprint", q))),
                eps_actual=as_num(raw.get(("eps_actual", q))),
                eps_surprise_pct=as_num(raw.get(("eps_surprise_pct", q))),
                eps_num_est=as_num(raw.get(("eps_num_est", q))),
                timing=known_timing.get(lab, timing_default), source="ciq",
            ))
        df = pd.DataFrame(recs)
        if len(df) >= 8:     # CapIQ is authoritative: drop other-source event rows for this ticker
            con.execute("DELETE FROM earnings WHERE ticker=? AND source<>'ciq'", (t,))
        n = db.upsert(con, "earnings", df, ["ticker", "fq_label"])
        # forward
        fwd = dict(
            ticker=t, asof=date.today().isoformat(),
            next_fq_label=labels.normalise(raw.get(("next_fq_label", 1)))
                or labels.from_period_end(as_date(raw.get(("next_period_end", 1))), fye),
            next_announce_date=(lambda d: d.strftime("%Y-%m-%d") if d is not None else None)(as_date(raw.get(("next_announce_date", 1)))),
            next_period_end=(lambda d: d.strftime("%Y-%m-%d") if d is not None else None)(as_date(raw.get(("next_period_end", 1)))),
            next_eps_est_now=as_num(raw.get(("next_eps_est_now", 1))),
            next_eps_est_m28d=as_num(raw.get(("next_eps_est_m28d", 1))),
            next_eps_num_est=as_num(raw.get(("next_eps_num_est", 1))),
            source="ciq",
        )
        prev = db.read_forward(con, t)
        for k in ("next_fq_label", "next_announce_date", "next_period_end", "next_eps_est_now",
                  "next_eps_est_m28d", "next_eps_num_est"):
            if fwd.get(k) is None and prev.get(k) is not None:
                fwd[k] = prev[k]                      # e.g. next call date not yet in CIQ -> keep Yahoo's
        if fwd["next_announce_date"] or fwd["next_eps_est_now"] is not None:
            db.upsert(con, "forward", pd.DataFrame([fwd]), ["ticker"])
        db.log(con, "ciq_pull.xlsx", "earnings", n, missing, t)
        status = "OK" if n >= layout["n_quarters"] - 2 else "CHECK - quarters missing (refresh again / see Test sheet)"
        print(f"  earnings {t:5s}: {n:2d} quarters, {missing} unresolved   {status}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(PULL_XLSX))
    ap.add_argument("--layout", default=str(PULL_LAYOUT))
    ap.add_argument("--use-fallback", action="store_true", help="read Fallback_Scalar sheet instead of Earnings")
    args = ap.parse_args()
    layout = json.loads(open(args.layout).read())
    print(f"Reading cached values from {args.xlsx}")
    wb = load_workbook(args.xlsx, data_only=True, read_only=False)
    con = db.connect()
    ingest_prices(wb["Prices"], layout["sheets"]["Prices"], con)
    if "Earnings" in layout["sheets"]:
        ingest_earnings(wb, layout, con, args.use_fallback)
    print("Done ->", db.DB_PATH)


if __name__ == "__main__":
    main()
